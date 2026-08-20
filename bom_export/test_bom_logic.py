#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""bom_export 纯逻辑单元测试（不依赖 CATIA）。

2026-08-13 重写: 老 gr_rules.json 系统已删除，规则推理走 V2 RuleSpec 引擎——
测试通过 m._engine_provider 注入内存规则集（RuleEngine 直接构造，不依赖规则文件）。

运行:  python -m pytest test_bom_logic.py -v
  或:  python test_bom_logic.py   (无 pytest 时跑内置 runner)
"""
import os
import sys

BASE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, BASE)
V2_DIR = os.path.normpath(os.path.join(BASE, "..", "V2"))
if V2_DIR not in sys.path:
    sys.path.insert(0, V2_DIR)

import bom_export as m
import v2_bridge

from rulespec.engine import RuleEngine
from rulespec.matcher import canonical_spec, extract_model_from_name


# ---------------- 内存测试规则集 ----------------

def _rule(rid, domain, scope, when, then, priority=500):
    return {"id": rid, "domain": domain, "priority": priority, "scope": scope,
            "when": when, "then": then,
            "meta": {"status": "active", "version": 1}}


_GR_SEQ = [0]


def _gr(kw, gr):
    _GR_SEQ[0] += 1
    return _rule(f"gr.test.{_GR_SEQ[0]:03d}", "gr", "part",
                 {"part.workingName": {"op": "contains", "value": kw}},
                 {"gr": gr})


TEST_RULES = [
    # ---- normalize: 同义词（反推杆→复位杆） ----
    _rule("normalize.test.001", "normalize", "part",
          {"part.workingName": {"op": "contains", "value": "反推杆"}},
          {"part.workingName": {"replaceAll": ["反推杆", "复位杆"]}}),
    # ---- gr ----
    _gr("模框", "模架"), _gr("复位杆", "模架"),
    _gr("热流道", "热流道"), _gr("吊环", "吊环"), _gr("隔水片", "隔水片"),
    _gr("镶块", "自制件"),
    # ---- material ----
    _rule("material.test.001", "material", "part",
          {"gr": {"op": "eq", "value": "模架"}},
          {"material": "50#锻件", "heatTreatment": ""}),
    _rule("material.test.002", "material", "part",
          {"gr": {"op": "eq", "value": "自制件"}},
          {"material": "SW2343"}),
    _rule("material.test.003", "material", "part",
          {"gr": {"op": "eq", "value": "隔水片"}},
          {"material": "黄铜"}),
    # ---- number: 分段 ----
    _rule("number.test.001", "number", "part",
          {"gr": {"op": "eq", "value": "模架"}},
          {"numberRange": {"min": 1, "max": 99}}),
    _rule("number.test.002", "number", "part",
          {"gr": {"op": "eq", "value": "自制件"}},
          {"numberRange": {"min": 100, "max": 199}}),
    _rule("number.test.003", "number", "global",
          {}, {"numberRange": {"min": 200, "max": 999}}, priority=100),
    # ---- companion: 零件级 + 策略 ----
    _rule("companion.test.001", "companion", "part",
          {"gr": {"op": "eq", "value": "模架"}},
          {"companions": [{"name": "螺钉", "spec": "M12*100", "qty": 4, "gr": ""}]}),
    _rule("companion.test.002", "companion", "part",
          {"gr": {"op": "eq", "value": "模架"}},
          {"companionGrPolicy": "follow-part"}),
    _rule("companion.test.003", "companion", "part",
          {"gr": {"op": "eq", "value": "自制件"}},
          {"companions": [{"name": "螺钉", "spec": "CB16-100", "qty": 2, "gr": ""},
                          {"name": "弹簧垫圈", "spec": "CBW16", "qty": 2, "gr": ""}]}),
    # ---- companion: 规格级（40*60*12 调整板） ----
    _rule("companion.test.004", "companion", "spec",
          {"part.workingName": {"op": "contains", "value": "调整板"},
           "spec.value": {"op": "eq", "value": "40*60*12"}},
          {"companions": [{"name": "螺钉", "spec": "CB8-16", "qty": 2,
                           "gr": "仓库备件"}]}),
    _rule("companion.test.005", "companion", "spec",
          {"part.workingName": {"op": "contains", "value": "调整板"},
           "spec.value": {"op": "eq", "value": "80*60*12"}},
          {"companions": [{"name": "螺钉", "spec": "CB8-16", "qty": 4,
                           "gr": "仓库备件"}]}),
]


def _make_engine():
    return {"engine": RuleEngine(TEST_RULES),
            "canonical_spec": canonical_spec,
            "extract_model_from_name": extract_model_from_name}


# 注入：bom_export 全部走内存引擎（隔离真实规则文件）。
# 推理链 infer_gr_and_detail → v2_bridge.infer_part → get_engine，
# 故只需替换 v2_bridge.get_engine 即可（2026-08-18 清理死注入点 _engine_provider）。
v2_bridge.get_engine = _make_engine


def _row(name, gr="小零件", qty=1, remark="", spec="", mat=""):
    return {"零件号": "", "零部件名": name, "数量": qty, "规格": spec,
            "材质": mat, "零件GR号": gr, "零部件GR名": "",
            "备注": remark, "加工备注": "", "_v2": {}}


def _mk(name, qty=1, remark="", spec="", mat=""):
    """经 V2 推理构造行（零件GR号/材质/备注/_v2 来自引擎输出）。"""
    d = m.infer_gr_and_detail(name)
    return {"零件号": "", "零部件名": name, "数量": qty, "规格": spec,
            "材质": mat or d["材质"], "零件GR号": d["零部件GR名"],
            "零部件GR名": "", "备注": remark, "加工备注": d["加工备注"],
            "_v2": d["_v2"]}


# ---------------- GR 推理 ----------------
def test_gr_basic():
    assert m.infer_gr_and_detail("定1模框")["零部件GR名"] == "模架"
    assert m.infer_gr_and_detail("热流道系统")["零部件GR名"] == "热流道"
    assert m.infer_gr_and_detail("吊环M12")["零部件GR名"] == "吊环"
    assert m.infer_gr_and_detail("隔水片")["零部件GR名"] == "隔水片"
    assert m.infer_gr_and_detail("定1镶块")["零部件GR名"] == "自制件"


def test_gr_default():
    # 未命中任何规则 → 默认 GR（DEFAULT_GR，原 default_gr=小零件）
    assert m.infer_gr_and_detail("未知名零件")["零部件GR名"] == "小零件"


def test_gr_material():
    assert m.infer_gr_and_detail("定1模框")["材质"] == "50#锻件"
    assert "SW2343" in m.infer_gr_and_detail("定1镶块")["材质"]
    assert m.infer_gr_and_detail("隔水片")["材质"] == "黄铜"


def test_gr_synonym():
    # V2 normalize 域同义词归一化: 反推杆→复位杆，再按复位杆命中模架
    r = m.infer_gr_and_detail("反推杆")
    assert r["零部件GR名"] == "模架"


# ---------------- 零件编号 ----------------
def test_assign_numbers():
    rows = [_mk("定1模框"),
            _mk("定1镶块"),
            _mk("调整板"),
            _row("螺钉", gr="仓库备件", remark="→ 定1模框")]
    out = m.assign_part_numbers(rows)
    by = {r["零部件名"]: r["零件号"] for r in out}
    assert by["定1模框"] == 1
    assert by["定1镶块"] == 100
    assert by["调整板"] == 200
    assert by["螺钉"] == 1  # 配套件继承父号


# ---------------- 配套补全 ----------------
def test_companions():
    rows = [_mk("定1模框", qty=2)]
    out = m.add_companions(rows)
    comps = [r for r in out if r["备注"].startswith("→ ")]
    assert len(comps) == 1
    # 模架的配套件 GR 跟随模架（V2 companionGrPolicy follow-part）
    assert all(r["零件GR号"] == "模架" for r in comps)


def test_companions_zizhi():
    """自制件补配套（CB16-100 螺钉 + CBW16 弹簧垫圈），非模架父件 → 仓库备件。"""
    rows = [_mk("定1镶块")]
    out = m.add_companions(rows)
    comps = [r for r in out if r["备注"].startswith("→ ")]
    assert len(comps) == 2
    names = sorted(c["零部件名"] for c in comps)
    assert names == ["弹簧垫圈", "螺钉"]
    assert all(c["零件GR号"] == "仓库备件" for c in comps)
    assert all(c["规格"] for c in comps)


def test_companions_v2_spec():
    """规格级配套：40*60*12 调整板 → CB8-16 螺钉 ×2（spec 作用域规则经 apply_v2_spec 生效）。"""
    rows = [_mk("调整板", spec="40*60*12")]
    m._apply_spec_gr_v2(rows)
    out = m.add_companions(rows)
    comps = [r for r in out if r["备注"].startswith("→ ")]
    assert len(comps) == 1
    assert comps[0]["数量"] == 2
    assert comps[0]["零件GR号"] == "仓库备件"


def test_companions_multi_spec():
    """多规格零件：紧固件数量 = Σ(规格实体数 × 该规格单件数)，同名同规格聚合（2026-08-13）。"""
    rows = [{"零件号": "", "零部件名": "调整板", "数量": 3,
             "规格": "", "材质": "", "零件GR号": "仓库备件", "零部件GR名": "",
             "备注": "", "加工备注": "", "_stp_path": "", "_v2": {},
             "_spec_list": ["40*60*12", "40*60*12", "80*60*12"]}]
    m._apply_spec_gr_v2(rows)
    out = m.add_companions(rows)
    comps = [r for r in out if r["备注"].startswith("→ ")]
    by = {(c["零部件名"], c["规格"]): c["数量"] for c in comps}
    # 40*60*12 ×2 个（每件 2 螺钉）+ 80*60*12 ×1 个（每件 4 螺钉）= 8
    assert by.get(("螺钉", "CB8-16")) == 8


def test_companions_structured_fields():
    """配套件带结构化字段 _is_companion/_parent_ref/_source（2026-08-19 P1-5 修复）。"""
    rows = [_mk("定1模框", qty=2)]
    rows[0]["_source"] = "A.CATPart"
    out = m.add_companions(rows)
    comps = [r for r in out if r.get("_is_companion")]
    assert len(comps) == 1
    assert comps[0]["_parent_ref"] == "定1模框"
    assert comps[0]["_source"] == "A.CATPart"
    assert comps[0]["备注"].startswith("→ ")  # 旧字符串契约保留兼容


def test_companion_over_limit_blocked():
    """P1-8：CB>M20 / CBW>16 配套件被拦截，不进入 BOM。"""
    rows = [{
        "零件号": "", "零部件名": "测试板", "数量": 1, "规格": "", "材质": "",
        "零件GR号": "自制件", "零部件GR名": "", "备注": "", "加工备注": "",
        "_v2": {"companions": [
            {"name": "螺钉", "spec": "CB24-100", "qty": 2, "gr": "标准件"},
            {"name": "弹簧垫圈", "spec": "CBW20", "qty": 2, "gr": "标准件"},
            {"name": "螺钉", "spec": "CB16-100", "qty": 2, "gr": "标准件"},
        ]},
    }]
    out = m.add_companions(rows)
    comps = [r for r in out if r.get("_is_companion")]
    # CB24 / CBW20 被拦截，CB16 保留
    assert [(c["零部件名"], c["规格"]) for c in comps] == [("螺钉", "CB16-100")]


def test_assign_numbers_batch_source_matching():
    """批量模式：同名父件按 _source 配对，配套件不串号（2026-08-19 修复）。"""
    rows = [
        {"零件号": "", "零部件名": "模框", "数量": 1, "规格": "", "材质": "",
         "零件GR号": "模架", "零部件GR名": "", "备注": "", "加工备注": "",
         "_v2": {"numberRange": {"min": 1, "max": 99}}, "_source": "A.CATPart"},
        {"零件号": "", "零部件名": "模框", "数量": 1, "规格": "", "材质": "",
         "零件GR号": "模架", "零部件GR名": "", "备注": "", "加工备注": "",
         "_v2": {"numberRange": {"min": 1, "max": 99}}, "_source": "B.CATPart"},
        {"零件号": "", "零部件名": "螺钉", "数量": 4, "规格": "M8*30", "材质": "",
         "零件GR号": "标准件", "零部件GR名": "", "备注": "→ 模框", "加工备注": "",
         "_v2": {}, "_source": "A.CATPart", "_is_companion": True, "_parent_ref": "模框"},
    ]
    out = m.assign_part_numbers(rows)
    by_name = {r["零部件名"]: r for r in out}
    # 两个同名主件各自有零件号（1、2）；A 文件的螺钉继承 A 文件的模框号
    nos = sorted(r["零件号"] for r in out if r["零部件名"] == "模框")
    assert nos == [1, 2]
    comp = by_name["螺钉"]
    main_a = [r for r in out if r["零部件名"] == "模框" and r["_source"] == "A.CATPart"][0]
    assert comp["零件号"] == main_a["零件号"] == 1


# ---------------- CATIA COM 兼容助手 ----------------
def test_as_part_document_fallback():
    """gen_py 静态绑定下 Document 无 Part → 转 PartDocument（2026-08-19 修复）。"""
    import bom_common
    import win32com.client as w

    class FakeDocNoPart:
        @property
        def Part(self):
            raise AttributeError("no Part")

    class FakeDocWithPart:
        Part = "PART"

    d = FakeDocWithPart()
    sentinel = object()
    orig_dispatch = w.dynamic.Dispatch

    w.dynamic.Dispatch = lambda doc: sentinel  # 动态绑定包装兜底
    try:
        assert bom_common.as_part_document(d) is d          # 已有 Part 直接返回原对象
        assert bom_common.as_part_document(FakeDocNoPart()) is sentinel  # 无 Part → 动态包装
    finally:
        w.dynamic.Dispatch = orig_dispatch


# ---------------- 自动更新 token ----------------
def test_updater_effective_token():
    """token 优先级：用户配置 > 构建期内嵌（2026-08-19 新增）。"""
    import updater
    import bom_token

    assert updater._effective_token({"token": "cfg-token"}) == "cfg-token"
    old = bom_token.EMBEDDED_TOKEN
    bom_token.EMBEDDED_TOKEN = "embedded-token"
    try:
        assert updater._effective_token({"token": ""}) == "embedded-token"
        assert updater._effective_token({}) == "embedded-token"
    finally:
        bom_token.EMBEDDED_TOKEN = old
    assert updater._effective_token({"token": ""}) == ""


def test_load_config_embedded_defaults():
    """配置全部内置：无任何配置文件也能拿到 repo/token/auto_check（2026-08-20）。"""
    import json
    import tempfile
    import updater

    assert updater.DEFAULT_REPO == "https://github.com/LittleDarkZero/moldbom"
    orig = updater.state_path
    try:
        d = tempfile.mkdtemp()
        updater.state_path = lambda: os.path.join(d, "state.json")
        cfg = updater.load_config()
        assert cfg["repo"] == updater.DEFAULT_REPO
        assert cfg["token"] == ""
        assert cfg["auto_check"] is True
        # 状态文件只影响 last_check/auto_check，绝不覆盖内置 repo/token
        with open(os.path.join(d, "state.json"), "w", encoding="utf-8") as f:
            json.dump({"repo": "https://evil.example/x", "token": "should-not-merge",
                       "auto_check": False, "last_check": "2026-08-20T00:00:00+00:00"}, f)
        cfg = updater.load_config()
        assert cfg["auto_check"] is False
        assert cfg["last_check"] == "2026-08-20T00:00:00+00:00"
        assert cfg["repo"] == updater.DEFAULT_REPO
        assert cfg["token"] == ""
    finally:
        updater.state_path = orig


def test_save_config_state_only():
    """save_config 只写运行时状态，绝不写 repo/token（2026-08-20）。"""
    import json
    import tempfile
    import updater

    orig = updater.state_path
    try:
        d = tempfile.mkdtemp()
        p = os.path.join(d, "state.json")
        updater.state_path = lambda: p
        updater.save_config({
            "repo": updater.DEFAULT_REPO,
            "token": "super-secret",
            "auto_check": False,
            "last_check": "2026-08-20T00:00:00+00:00",
        })
        with open(p, "r", encoding="utf-8") as f:
            state = json.load(f)
        assert set(state) == {"auto_check", "last_check"}
        assert "token" not in state and "repo" not in state
    finally:
        updater.state_path = orig

def test_resolve_download_url():
    """私有仓库优先 API asset 地址；无 token 用浏览器地址（2026-08-20 新增）。"""
    import updater
    import bom_token

    info = {"url": "https://github.com/x/y/releases/download/v1/a.exe",
            "api_url": "https://api.github.com/repos/x/y/releases/assets/123"}
    # 有 token → api_url
    assert updater._resolve_download_url(info, {"token": "t"}) == info["api_url"]
    # 无 token → 浏览器 url
    assert updater._resolve_download_url(info, {"token": ""}) == info["url"]
    # 无 api_url 字段 → url
    assert updater._resolve_download_url({"url": info["url"]}, {"token": "t"}) == info["url"]
    # 内嵌 token 同样走 api_url
    old = bom_token.EMBEDDED_TOKEN
    bom_token.EMBEDDED_TOKEN = "embedded"
    try:
        assert updater._resolve_download_url(info, {"token": ""}) == info["api_url"]
    finally:
        bom_token.EMBEDDED_TOKEN = old

def test_manifest_urls_order():
    """update.json 候选源顺序：镜像 → api.github.com → raw（2026-08-20）。"""
    import updater

    cfg = {"repo": updater.DEFAULT_REPO, "mirror": ""}
    urls = updater._manifest_urls(cfg)
    assert urls[0].startswith("https://api.github.com/repos/LittleDarkZero/moldbom/contents/update.json")
    assert urls[1].startswith("https://raw.githubusercontent.com/LittleDarkZero/moldbom/main/update.json")
    assert len(urls) == 2

    cfg["mirror"] = "https://gitee.com/x/y/raw/main"
    urls = updater._manifest_urls(cfg)
    assert urls[0] == "https://gitee.com/x/y/raw/main/update.json"
    assert urls[1].startswith("https://api.github.com/")
    assert urls[2].startswith("https://raw.githubusercontent.com/")

    try:
        updater._manifest_urls({"repo": "", "mirror": ""})
        assert False, "应抛出 UpdaterError"
    except updater.UpdaterError:
        pass


def test_fetch_manifest_fallback():
    """fetch_manifest 多源回退：api 失败 → raw 成功；镜像命中后不再请求 GitHub（2026-08-20）。"""
    import updater

    class _FakeResp:
        def __init__(self, content):
            self._content = content

        def read(self):
            return self._content

        def close(self):
            pass

    calls = []

    def fake_get(url, cfg, headers=None):
        calls.append(url)
        if url.startswith("https://api.github.com/"):
            raise updater.UpdaterError("api down")
        return _FakeResp(b'{"exe": {"version": "9.3.2"}}')

    orig = updater._http_get
    updater._http_get = fake_get
    try:
        data = updater.fetch_manifest({"repo": updater.DEFAULT_REPO, "mirror": ""})
        assert data["exe"]["version"] == "9.3.2"
        assert calls[0].startswith("https://api.github.com/")
        assert calls[1].startswith("https://raw.githubusercontent.com/")
    finally:
        updater._http_get = orig

    calls.clear()

    def fake_get2(url, cfg, headers=None):
        calls.append(url)
        return _FakeResp(b'{"exe": {"version": "9.3.3"}}')

    updater._http_get = fake_get2
    try:
        data = updater.fetch_manifest({"repo": updater.DEFAULT_REPO,
                                       "mirror": "https://gitee.com/x/y/raw/main"})
        assert data["exe"]["version"] == "9.3.3"
        assert calls == ["https://gitee.com/x/y/raw/main/update.json"]
    finally:
        updater._http_get = orig



# ---------------- 输出格式 ----------------
def test_write_csv(tmp_path=None):
    import tempfile
    d = tempfile.mkdtemp()
    p = m.write_bom([_row("定1模框", gr="模架")], os.path.join(d, "o"), fmt="csv")
    assert p.endswith(".csv") and os.path.exists(p)
    with open(p, encoding="utf-8-sig") as f:
        head = f.readline()
    assert "零部件名" in head


def test_write_xlsx():
    import tempfile
    d = tempfile.mkdtemp()
    p = m.write_bom([_row("定1模框", gr="模架")], os.path.join(d, "o"), fmt="xlsx")
    assert p.endswith(".xlsx") and os.path.exists(p)


# ---------------- 按 GR 分组 / 拆分 BOM（2026-08-13 新功能1） ----------------
def test_group_by_gr():
    rows = [_row("定1模框", gr="模架", qty=1),
            _row("螺钉", gr="模架", remark="→ 定1模框"),
            _row("定1镶块", gr="自制件", qty=1),
            _row("调整板", gr="仓库备件", qty=2),
            _row("螺钉", gr="标准件", remark="→ 调整板")]
    groups = m._group_by_gr(rows)
    assert set(groups.keys()) == {"模架", "自制件", "仓库备件", "标准件"}
    # 紧固件按自己 GR 归组（螺钉 GR=标准件，不跟随父件"调整板"的仓库备件）
    assert [r["零部件名"] for r in groups["模架"]] == ["定1模框", "螺钉"]
    assert [r["零部件名"] for r in groups["仓库备件"]] == ["调整板"]
    assert [r["零部件名"] for r in groups["标准件"]] == ["螺钉"]
    assert [r["零部件名"] for r in groups["自制件"]] == ["定1镶块"]


def test_group_by_gr_unclassified():
    # 无 GR 的零件归"未分类"
    groups = m._group_by_gr([_row("无名件", gr="")])
    assert "未分类" in groups


def test_write_bom_by_gr():
    import tempfile
    import openpyxl
    d = tempfile.mkdtemp()
    rows = [_row("定1模框", gr="模架", qty=1),
            _row("螺钉", gr="模架", remark="→ 定1模框"),
            _row("定1镶块", gr="自制件", qty=1)]
    out = os.path.join(d, "26-1-99-BOM.xlsx")
    m.write_bom(rows, out)
    paths = m.write_bom_by_gr(rows, out, mold_num="26-1-99")
    names = sorted(os.path.basename(p) for p in paths)
    assert names == ["26-1-99-模架.xlsx", "26-1-99-自制件.xlsx"]
    wb = openpyxl.load_workbook(os.path.join(d, "26-1-99-模架.xlsx"))
    ws = wb.active
    parts = [ws.cell(row=r, column=2).value for r in range(3, ws.max_row + 1)]
    assert parts == ["定1模框", "螺钉"]


def test_safe_name():
    assert m._safe_name('a/b\\c:d*e?f"g<h>i|j') == 'a_b_c_d_e_f_g_h_i_j'


# ---------------- Pipeline ----------------
def test_pipeline_register():
    n0 = len(m._PIPELINE_STAGES)

    @m.stage("测试阶段-x", position=1)
    def _s(ctx):
        ctx["_t"] = 1
        return ctx
    assert len(m._PIPELINE_STAGES) == n0 + 1
    # 清理，避免污染其他用例
    m._PIPELINE_STAGES.pop(1)


def test_pipeline_subset():
    ctx = {"results": [_mk("定1模框")]}
    sub = [(n, f) for n, f in m._PIPELINE_STAGES if n in ("配套补全", "零件编号")]
    out = m.run_pipeline(ctx, stages=sub)
    assert "_timings" in out
    assert out["results"][0]["零件号"] == 1


# ---------------- 规格测量引擎 geometry_engine ----------------
# 纯几何、无 CATIA 依赖。生成器为确定性随机（RandomState 42），断言值稳定。

def test_geometry_box():
    ge = m.geometry_engine
    pts = ge.generate_box_points(size=(640, 218, 160), n_points=3000,
                                 noise=0.05, rotation_deg=(20, 30, 0))
    a = ge.analyze_points(pts, name="测试长方体")
    assert a["shape_en"] == "box"
    assert ge.format_spec(a) == "640.3*218.3*160.3"  # 保留 1 位小数（整数去 .0）


def test_geometry_cylinder():
    ge = m.geometry_engine
    pts = ge.generate_cylinder_points(radius=35, height=310, n_points=3000,
                                      noise=0.04, rotation_deg=(30, 0, 45))
    a = ge.analyze_points(pts, name="测试圆柱")
    assert a["shape_en"] == "cylinder"
    spec = ge.format_spec(a)
    assert spec.startswith("Φ") and "×310" in spec  # Φ74×310（噪声使直径略大）


def test_geometry_disk_phi_format():
    ge = m.geometry_engine
    pts = ge.generate_cylinder_points(radius=80, height=25, n_points=3000,
                                      noise=0.04, rotation_deg=(0, 60, 0))
    a = ge.analyze_points(pts, name="测试圆盘")
    assert a["shape_en"] == "cylinder"  # 圆盘也走 Φ 格式
    assert ge.format_spec(a).startswith("Φ16")


def test_geometry_volume_and_degenerate():
    ge = m.geometry_engine
    assert ge.analyze_points([[0, 0, 0], [1, 1, 1]], name="退化") is None
    pts = ge.generate_box_points(size=(100, 50, 30), n_points=2000)
    a = ge.analyze_points(pts, name="小长方体")
    assert ge.analysis_volume(a) > 0


def test_projection_classification():
    """2026-08-11 投影特征分类：box=3矩形 / cylinder=1圆+2等宽矩形 / 正方体≠球。"""
    ge = m.geometry_engine
    a = ge.analyze_points(
        ge.generate_box_points(size=(100, 60, 20), n_points=1500,
                               noise=0.02, rotation_deg=(20, 30, 0)),
        name="长方体")
    assert a["shape_en"] == "box", (a["shape_en"], a["decision"])
    c = ge.analyze_points(
        ge.generate_cylinder_points(radius=20, height=80, n_points=1500,
                                    noise=0.02, rotation_deg=(10, 20, 0)),
        name="长圆柱")
    assert c["shape_en"] == "cylinder"
    assert ge.format_spec(c).startswith("Φ4"), ge.format_spec(c)   # Φ40.x
    d = ge.analyze_points(
        ge.generate_cylinder_points(radius=80, height=25, n_points=3000,
                                    noise=0.02, rotation_deg=(10, 15, 0)),
        name="圆盘")
    assert d["shape_en"] == "cylinder"
    assert ge.format_spec(d).startswith("Φ16"), ge.format_spec(d)  # Φ160.x（非 Φ26）
    b = ge.analyze_points(
        ge.generate_box_points(size=(40, 40, 40), n_points=1500, noise=0.02),
        name="正方体")
    assert b["shape_en"] == "box", (b["shape_en"], b["decision"])


# ---------------- 内置 runner ----------------
def _run_all():
    fns = [(k, v) for k, v in globals().items()
           if k.startswith("test_") and callable(v)]
    passed = failed = 0
    for name, fn in fns:
        try:
            fn()
            print(f"  PASS {name}")
            passed += 1
        except AssertionError as e:
            print(f"  FAIL {name}: {e}")
            failed += 1
        except Exception as e:
            print(f"  ERR  {name}: {type(e).__name__}: {e}")
            failed += 1
    print(f"\n{passed} passed, {failed} failed")
    return failed


if __name__ == "__main__":
    sys.exit(1 if _run_all() else 0)
