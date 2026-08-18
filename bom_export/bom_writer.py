# -*- coding: utf-8 -*-
"""BomExport Excel/CSV 输出模块（2026-08-18 重构自 bom_export.py 模块7 下半）。

统一 BOM 写出入口（write_bom），对齐模具明细表模板；支持按 GR 拆分成
多份明细表（write_bom_by_gr）。
"""

import os

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

from bom_common import log
from bom_utils import _safe_name, _group_by_gr

HEADER_FONT = Font(name="微软雅黑", bold=True, size=10)
HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_FONT = Font(name="微软雅黑", size=9)
CELL_ALIGNMENT = Alignment(vertical="center", wrap_text=True)
CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

BOM_COLUMNS = ["零件号", "零部件名", "数量", "规格", "材质", "零件GR号", "零部件GR名", "备注", "加工备注"]


def write_bom_excel(data_list: list, output_path: str, module_name: str = ""):
    wb = Workbook(); ws = wb.active
    ws.title = "明细表1" if not module_name else module_name
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(BOM_COLUMNS))
    title_cell = ws.cell(row=1, column=1, value="模具明细表")
    title_cell.font = Font(name="微软雅黑", bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    for col_idx, col_name in enumerate(BOM_COLUMNS, 1):
        cell = ws.cell(row=2, column=col_idx, value=col_name)
        cell.font = HEADER_FONT; cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT; cell.border = THIN_BORDER

    for row_idx, item in enumerate(data_list, 3):
        for col_idx, col_name in enumerate(BOM_COLUMNS, 1):
            value = item.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = CELL_FONT; cell.border = THIN_BORDER
            if col_name in ("数量", "零件号"):
                cell.alignment = CENTER_ALIGNMENT
            else:
                cell.alignment = CELL_ALIGNMENT

    col_widths = [10, 28, 10, 18, 26, 14, 14, 24, 40]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=2, column=col_idx).column_letter].width = width
    ws.freeze_panes = "A3"
    wb.save(output_path)
    return output_path


def write_bom_csv(data_list: list, output_path: str):
    """导出 BOM 为 UTF-8-SIG CSV（Excel 可直接打开，中文不乱码）。验证输出格式扩展性。"""
    import csv
    with open(output_path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(BOM_COLUMNS)
        for item in data_list:
            w.writerow([item.get(c, "") for c in BOM_COLUMNS])
    return output_path


def write_bom(data_list: list, output_path: str, fmt: str = "xlsx", module_name: str = ""):
    """统一输出入口：按 fmt 分发到具体 writer。新增格式只需在此注册。"""
    fmt = (fmt or "xlsx").lower()
    if fmt == "csv":
        if not output_path.lower().endswith(".csv"):
            output_path = os.path.splitext(output_path)[0] + ".csv"
        return write_bom_csv(data_list, output_path)
    # 默认 xlsx
    if not output_path.lower().endswith(".xlsx"):
        output_path = os.path.splitext(output_path)[0] + ".xlsx"
    return write_bom_excel(data_list, output_path, module_name)


def write_bom_by_gr(data_list: list, output_path: str, mold_num: str, fmt: str = "xlsx"):
    """按 GR 拆分成多个 BOM 文件（2026-08-13 新功能1），与总 BOM 同目录。

    命名: {模号}-{GR}.xlsx（主件与紧固件都按各自 GR 名归组）。
    返回写入的文件路径列表。
    """
    groups = _group_by_gr(data_list)
    out_dir = os.path.dirname(os.path.abspath(output_path))
    written = []
    for gr, rows in groups.items():
        gr_safe = _safe_name(gr)
        path = os.path.join(out_dir, f"{_safe_name(mold_num + '-' + gr)}.{fmt}")
        write_bom(rows, path, fmt=fmt, module_name=gr_safe)
        written.append(path)
    log.info("按 GR 拆分 BOM: %d 份（%s）", len(written),
             ", ".join(_safe_name(g) for g in groups))
    return written
